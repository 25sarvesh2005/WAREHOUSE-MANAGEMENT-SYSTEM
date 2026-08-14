"""
--------------------------------------------------------------------------------
File        : core/services/import_export/opening_inventory_parser.py
Purpose     : Parse opening inventory CSV/XLSX files into staged row dictionaries.

Responsibilities:
    - Parse supported spreadsheet formats without mutating inventory.
    - Preserve raw row evidence, workbook/sheet names, row numbers, and hashes.
    - Enforce explicit source-column expectations for migration staging.

Flow:
    Upload/CLI file bytes -> parser -> MigrationController.submit_staged_rows()

Used By:
    - core/controllers/migration_controller.py
    - tools/import_opening_inventory.py

Returns:
    list[dict[str, object]] - Staged row dictionaries compatible with migration input.

Raises:
    ValueError: If file type, dependency, workbook, or column layout is invalid.
--------------------------------------------------------------------------------
"""

from __future__ import annotations

import csv
import hashlib
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterable

from common.logger import get_logger

logger = get_logger(__name__)

REQUIRED_COLUMNS = {
    "seller_code",
    "sku",
    "warehouse_code",
    "inventory_state",
    "quantity",
}
OPTIONAL_COLUMNS = {"upc", "location_code"}
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def compute_opening_inventory_row_hash(item: dict[str, Any]) -> str:
    """
    Compute a deterministic SHA-256 hash for raw opening inventory row content.

    Args:
        item: Raw staged row dictionary.

    Returns:
        str: SHA-256 hex digest.
    """
    raw_str = (
        f"{item.get('source_workbook', '')}|"
        f"{item.get('source_sheet', '')}|"
        f"{item.get('source_row_number', '')}|"
        f"{item.get('raw_seller_code', '') or ''}|"
        f"{item.get('raw_sku', '') or ''}|"
        f"{item.get('raw_upc', '') or ''}|"
        f"{item.get('raw_warehouse_code', '') or ''}|"
        f"{item.get('raw_location_code', '') or ''}|"
        f"{item.get('raw_inventory_state', '') or ''}|"
        f"{item.get('raw_quantity', '') or ''}"
    )
    return hashlib.sha256(raw_str.encode("utf-8")).hexdigest()


def parse_opening_inventory_file(
    file_name: str,
    file_bytes: bytes,
) -> list[dict[str, Any]]:
    """
    Parse a CSV or XLSX opening inventory file into staged row dictionaries.

    Args:
        file_name: Original uploaded/imported filename.
        file_bytes: File contents.

    Returns:
        list[dict[str, Any]]: Parsed staged row dictionaries.

    Raises:
        ValueError: If the extension or source layout is unsupported.
    """
    extension = Path(file_name).suffix.lower()
    if extension == ".csv":
        return parse_opening_inventory_csv(file_name, file_bytes)
    if extension == ".xlsx":
        return parse_opening_inventory_xlsx(file_name, file_bytes)
    raise ValueError("Unsupported opening inventory file type. Use .csv or .xlsx.")


def parse_opening_inventory_csv(
    file_name: str,
    file_bytes: bytes,
) -> list[dict[str, Any]]:
    """
    Parse a CSV opening inventory file.

    Args:
        file_name: Source filename.
        file_bytes: CSV bytes, expected as UTF-8 or UTF-8 with BOM.

    Returns:
        list[dict[str, Any]]: Parsed staged row dictionaries.

    Raises:
        ValueError: If decoding or required columns fail.
    """
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ValueError("CSV file must be UTF-8 encoded.") from error

    reader = csv.reader(StringIO(text))
    try:
        header = next(reader)
    except StopIteration as error:
        raise ValueError("Opening inventory CSV is empty.") from error

    columns = _build_column_index(header)
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(reader, start=2):
        if _row_is_empty(values):
            continue
        rows.append(_build_staged_row(file_name, "CSV", row_number, columns, values))

    logger.info("Parsed %s staged rows from CSV %s", len(rows), file_name)
    return rows


def parse_opening_inventory_xlsx(
    file_name: str,
    file_bytes: bytes,
) -> list[dict[str, Any]]:
    """
    Parse an XLSX opening inventory workbook.

    Args:
        file_name: Source filename.
        file_bytes: XLSX bytes.

    Returns:
        list[dict[str, Any]]: Parsed staged row dictionaries from all sheets.

    Raises:
        ValueError: If openpyxl is unavailable or workbook layout is invalid.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError(
            "XLSX import requires openpyxl. Install requirements.txt before use."
        ) from error

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as error:
        raise ValueError("Unable to read opening inventory XLSX workbook.") from error

    parsed_rows: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            sheet_rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(sheet_rows)
            except StopIteration:
                continue
            columns = _build_column_index([_raw_cell_to_string(value) for value in header])

            for row_number, values in enumerate(sheet_rows, start=2):
                row_values = [_raw_cell_to_string(value) for value in values]
                if _row_is_empty(row_values):
                    continue
                parsed_rows.append(
                    _build_staged_row(
                        file_name,
                        worksheet.title,
                        row_number,
                        columns,
                        row_values,
                    )
                )
    finally:
        workbook.close()

    logger.info("Parsed %s staged rows from XLSX %s", len(parsed_rows), file_name)
    return parsed_rows


def _build_column_index(header: Iterable[str]) -> dict[str, int]:
    """
    Build a normalized column-name to index map.

    Args:
        header: Header row values.

    Returns:
        dict[str, int]: Explicit supported column map.

    Raises:
        ValueError: If required columns are missing or duplicate headers exist.
    """
    column_index: dict[str, int] = {}
    for index, raw_name in enumerate(header):
        normalized = _normalize_column_name(raw_name)
        if not normalized:
            continue
        if normalized in column_index:
            raise ValueError(f"Duplicate source column '{normalized}' is not allowed.")
        column_index[normalized] = index

    missing = sorted(REQUIRED_COLUMNS - set(column_index))
    if missing:
        raise ValueError(f"Missing required opening inventory columns: {missing}.")
    return column_index


def _normalize_column_name(value: str | None) -> str:
    """
    Normalize an explicit supported column name.

    Args:
        value: Raw header value.

    Returns:
        str: Normalized column name, or empty string for blank headers.
    """
    raw = (value or "").strip().lower()
    normalized = raw.replace(" ", "_").replace("-", "_")
    allowed_columns = REQUIRED_COLUMNS | OPTIONAL_COLUMNS
    if normalized in allowed_columns:
        return normalized
    return normalized


def _row_is_empty(values: Iterable[object]) -> bool:
    """
    Return whether every cell in a row is blank.

    Args:
        values: Raw row cell values.

    Returns:
        bool: True when the row contains no visible values.
    """
    return all(_raw_cell_to_string(value) == "" for value in values)


def _raw_cell_to_string(value: object) -> str:
    """
    Convert a spreadsheet cell value to preserved raw text.

    Args:
        value: Raw CSV/XLSX cell value.

    Returns:
        str: Trimmed raw string value.
    """
    if value is None:
        return ""
    return str(value).strip()


def _value_at(values: list[object], columns: dict[str, int], column_name: str) -> str | None:
    """
    Read a raw string value by column name.

    Args:
        values: Row values.
        columns: Column index map.
        column_name: Target column name.

    Returns:
        str | None: Raw text value or None when the optional column is absent/blank.
    """
    index = columns.get(column_name)
    if index is None or index >= len(values):
        return None
    raw = _raw_cell_to_string(values[index])
    return raw or None


def _build_staged_row(
    workbook_name: str,
    sheet_name: str,
    row_number: int,
    columns: dict[str, int],
    values: Iterable[object],
) -> dict[str, Any]:
    """
    Build a staged row dictionary from source row values.

    Args:
        workbook_name: Source workbook/file name.
        sheet_name: Source sheet name.
        row_number: Original source row number.
        columns: Column index map.
        values: Source row values.

    Returns:
        dict[str, Any]: Migration staged row payload.
    """
    row_values = list(values)
    staged_row: dict[str, Any] = {
        "source_workbook": Path(workbook_name).name,
        "source_sheet": sheet_name,
        "source_row_number": row_number,
        "raw_seller_code": _value_at(row_values, columns, "seller_code"),
        "raw_sku": _value_at(row_values, columns, "sku"),
        "raw_upc": _value_at(row_values, columns, "upc"),
        "raw_warehouse_code": _value_at(row_values, columns, "warehouse_code"),
        "raw_location_code": _value_at(row_values, columns, "location_code"),
        "raw_inventory_state": _value_at(row_values, columns, "inventory_state"),
        "raw_quantity": _value_at(row_values, columns, "quantity"),
    }
    staged_row["source_hash"] = compute_opening_inventory_row_hash(staged_row)
    return staged_row
